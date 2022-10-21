"""   /api/v1/product/get_products route   """
from falcon import Request, Response
from spectree import Response as resp

from src.schemas.base import Base401, Base500, BaseHeader
from src.schemas.product import GetDataProduct, GetProduct200, Product_tag
from src.sql.connection import async_session, engine
# from src.utils import api
import openpyxl
from awaits.awaitable import awaitable
from awaits.shoot import shoot
import logging
from src.schemas.product import UpdateDataOffer, UpdateOffer200
from src.methods.product import async_update_product
from fastapi import APIRouter, Body, Depends
from src.middleware import process_resource
from src.methods.product import async_get_product

router = APIRouter()


# @shoot
# def excel(out):
#     wb = openpyxl.Workbook()
#     sheet = wb.active
#     i = 1
#     for key1 in out['data'][0].keys():
#         for key2 in out['data'][0][key1].keys():
#             sheet.cell(row=1, column=i).value = f'{key1}_{key2}'
#             i += 1
#     j = 1
#     for item1 in out['data']:
#         i = 1
#         j += 1
#         for item2 in item1.values():
#             for item3 in item2.values():
#                 sheet.cell(row=j, column=i).value = item3
#                 i += 1
#     wb.active.title = 'Предзаказ Яндекс'
#     wb.close()
#     wb.save("Предзаказ.xlsx")


@router.post("/api/v1/product/get_products", tags=["Product"], response_model=GetProduct200)
async def update_offer(body: GetDataProduct, JWT: bool = Depends(process_resource)):
    out = await async_get_product(body, JWT['user_id'])
    return out

# if "download_to_excel" in data:
#     if data['download_to_excel']:
#         import asyncio
#
#         excel(out)
